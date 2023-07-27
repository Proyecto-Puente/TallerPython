# Ejercicio 4
# Taller: introducción a python
# Día 9

# Consigna:
#   Leer datos desde un archivo XLSX y realizar operaciones matemáticas
#   con ellos (por ejemplo, multiplicar una columna por un valor
#   constante).


import openpyxl

archivo = "ejemplo.xlsx"

libro = openpyxl.load_workbook(archivo)
hoja = libro.active
print(libro.sheetnames, hoja.title)

for nro, fila in enumerate(hoja.iter_rows()):
    if nro != 0:
        print(fila[4].value/2)
        hoja["F"+str(nro)].value = fila[4].value/2
    else:
        # La fila contiene el encabezado.
        print(fila[4].value)
libro.save(archivo)