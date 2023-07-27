# Ejercicio 5
# Taller: introducción a python
# Día 9

# Consigna:
#   Crear una aplicación que permita al usuario ingresar datos y guardarlos
#   en un archivo CSV o XLSX, según su elección.

import csv
from os.path import exists
from openpyxl import Workbook, load_workbook

def guardar_datos(datos, formato = "csv", ruta_csv = "datos.csv", ruta_xlsx = "datos.xlsx"):
    if formato == "csv":
        with open(ruta_csv, "a", newline="") as archivo:
            escritor = csv.writer(archivo)
            escritor.writerow(datos)
    
    elif formato == "xlsx":
        if exists(ruta_xlsx):
            libro = load_workbook(ruta_xlsx)
        else:
            libro = Workbook()
        
        hoja = libro.active
        hoja.append(datos)
        libro.save(ruta_xlsx)

while True:
    datos = input("Ingrese los datos separados por comas (nombre, apellido, edad): ")
    datos = [campo.strip() for campo in datos.split(",")]
    
    formato = input("Ingrese el formato de archivo (csv o xlsx): ")
    guardar_datos(datos, formato.strip().lower())

    continuar = input("¿Desea ingresar más datos? (s/n): ")
    if continuar.lower() != "s":
        break
